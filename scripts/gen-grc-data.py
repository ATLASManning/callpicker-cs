"""
Regenera app/churn/aaa-grc-data.ts desde el Excel de Zoho Analytics.

Uso:
    python scripts/gen-grc-data.py "D:/Archivos/GRC_AAA_2026.xlsx"
    python scripts/gen-grc-data.py "<ruta>" ENE FEB MAR ABR MAY JUN JUL AGO

Sin lista de meses toma ENE..AGO (el corte vigente). Para incorporar un mes
nuevo basta con añadir su clave al final de la llamada.

Acepta los DOS formatos que ha entregado Zoho Analytics:
  (a) una hoja por mes, nombrada ENE, FEB, … (DT_Churn_etiquetas.xlsx)
  (b) una sola hoja con todos los meses, separados por la columna
      "Mes Nombre"                                (GRC_AAA_2026.xlsx)
El formato se detecta solo: si las hojas por mes no existen, se agrupa por
la columna 0. Un mes pedido que no aparezca en los datos aborta la
generación en vez de escribir un archivo incompleto en silencio.

Columnas esperadas, en este orden:
  0 Mes Nombre · 1 Cliente · 2 clasificacion_cliente · 3 Año_Mes fecha
  4 Facturas_2026 · 5 Meses Activo · 6 Importe Acumulado Recurrente
  7 MRR Inicio Contrato · 8 MRR Fin Contrato · 9 Ingreso Ganado Contrato
  10 Movimiento MRR · 11 Ingreso Perdido Real
  12 Ingreso Perdido Fraude-Reestructura · 13 Rango MRR Fin Contrato

Requiere openpyxl:  python -m pip install openpyxl
"""
import sys
import os
import openpyxl

NOMBRES = {
    'ENE': 'Enero', 'FEB': 'Febrero', 'MAR': 'Marzo', 'ABR': 'Abril',
    'MAY': 'Mayo', 'JUN': 'Junio', 'JUL': 'Julio', 'AGO': 'Agosto',
    'SEP': 'Septiembre', 'OCT': 'Octubre', 'NOV': 'Noviembre', 'DIC': 'Diciembre',
}
POR_OMISION = ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO']

BS = chr(92)
QT = chr(39)


def esc(valor):
    texto = str(valor or '').strip()
    return texto.replace(BS, BS + BS).replace(QT, BS + QT)


def num(valor):
    try:
        return round(float(valor or 0), 2)
    except Exception:
        return 0


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(1)

    origen = sys.argv[1]
    hojas = sys.argv[2:] or POR_OMISION
    destino = os.path.join(os.path.dirname(__file__), '..', 'app', 'churn', 'aaa-grc-data.ts')
    destino = os.path.normpath(destino)

    wb = openpyxl.load_workbook(origen, read_only=True, data_only=True)

    def fila(r):
        return {
            'cliente': esc(r[1]), 'clas': esc(r[2]),
            'facturas': int(num(r[4])), 'meses': int(num(r[5])),
            'acumulado': num(r[6]), 'mrrInicio': num(r[7]), 'mrrFin': num(r[8]),
            'ganado': num(r[9]), 'movimiento': esc(r[10]),
            'perdido': num(r[11]), 'perdido2': num(r[12]), 'rango': esc(r[13]),
        }

    porMes = {}
    faltan = [h for h in hojas if h not in wb.sheetnames]
    if not faltan:
        # (a) una hoja por mes
        formato = 'una hoja por mes'
        for clave in hojas:
            it = wb[clave].iter_rows(values_only=True)
            next(it)
            porMes[NOMBRES.get(clave, clave)] = [fila(r) for r in it if r[1]]
    else:
        # (b) una sola hoja, separada por la columna "Mes Nombre"
        formato = 'hoja unica con columna de mes'
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        next(it)
        for r in it:
            if not r[1] or not r[0]:
                continue
            porMes.setdefault(esc(r[0]), []).append(fila(r))
        pedidos = [NOMBRES.get(h, h) for h in hojas]
        ausentes = [m for m in pedidos if m not in porMes]
        if ausentes:
            print('Estos meses no aparecen en la columna "Mes Nombre": %s'
                  % ', '.join(ausentes))
            print('Disponibles: %s' % ', '.join(porMes.keys()))
            raise SystemExit(1)
        sobran = [m for m in porMes if m not in pedidos]
        if sobran:
            print('AVISO: el archivo trae meses que no se pidieron y quedan FUERA: %s'
                  % ', '.join(sobran))

    meses = []
    total = 0
    for clave in hojas:
        nombre = NOMBRES.get(clave, clave)
        filas = porMes[nombre]
        filas.sort(key=lambda x: -(x['perdido'] + x['perdido2']))
        total += len(filas)
        meses.append((nombre, filas))

    rango = '%s a %s 2026' % (meses[0][0], meses[-1][0])
    L = []
    A = L.append
    A('/* ═══════════════════════════════════════════════════════════════════════')
    A('   GRC · CLIENTES POR CLASIFICACIÓN — %s' % rango)
    A('   Fuente: %s (Zoho Analytics) — %s.' % (os.path.basename(origen), formato))
    A('   Se cargan TODAS las clasificaciones (AAA · AA · A · B · C) para poder')
    A('   filtrarlas en la vista; el módulo resalta AAA por omisión.')
    A('')
    A('   GENERADO — no editar a mano. Regenerar con:')
    A('     python scripts/gen-grc-data.py "<ruta del Excel>" %s' % ' '.join(hojas))
    A('═══════════════════════════════════════════════════════════════════════ */')
    A('')
    A('export interface AAAGrcRow {')
    A('  cliente: string')
    A('  /** AAA · AA · A · B · C */')
    A('  clas: string')
    A('  facturas: number')
    A('  meses: number')
    A('  acumulado: number')
    A('  mrrInicio: number')
    A('  mrrFin: number')
    A('  ganado: number')
    A('  movimiento: string')
    A('  /** Ingreso perdido real */')
    A('  perdido: number')
    A('  /** Ingreso perdido por fraude o reestructura */')
    A('  perdido2: number')
    A('  /** Rango de MRR al cierre del contrato */')
    A('  rango: string')
    A('}')
    A('')
    A('export interface AAAGrcMes {')
    A('  mes: string')
    A('  clientes: AAAGrcRow[]')
    A('}')
    A('')
    A('export const AAA_GRC_2026: AAAGrcMes[] = [')
    for nombre, filas in meses:
        A('  {')
        A("    mes: '%s'," % nombre)
        A('    clientes: [')
        for f in filas:
            A("      { cliente: '%s', clas: '%s', facturas: %d, meses: %d, acumulado: %s, "
              "mrrInicio: %s, mrrFin: %s, ganado: %s, movimiento: '%s', perdido: %s, "
              "perdido2: %s, rango: '%s' }," % (
                  f['cliente'], f['clas'], f['facturas'], f['meses'], f['acumulado'],
                  f['mrrInicio'], f['mrrFin'], f['ganado'], f['movimiento'],
                  f['perdido'], f['perdido2'], f['rango']))
        A('    ],')
        A('  },')
    A(']')
    A('')
    A('/** Todas las filas aplanadas, con el mes incorporado. */')
    A('export const AAA_GRC_FLAT: Array<AAAGrcRow & { mes: string }> =')
    A('  AAA_GRC_2026.flatMap(m => m.clientes.map(c => ({ ...c, mes: m.mes })))')
    A('')

    with open(destino, 'w', encoding='utf-8', newline='\n') as fh:
        fh.write('\n'.join(L))

    perdido = sum(f['perdido'] for _, fs in meses for f in fs)
    fraude = sum(f['perdido2'] for _, fs in meses for f in fs)
    print('Escrito: %s' % destino)
    print('%d filas en %d meses' % (total, len(meses)))
    for nombre, filas in meses:
        aaa = sum(1 for f in filas if f['clas'] == 'AAA')
        print('  %-11s %4d filas · %3d AAA' % (nombre, len(filas), aaa))
    print('Ingreso perdido real:   $%s' % format(perdido, ',.2f'))
    print('Fraude / reestructura:  $%s' % format(fraude, ',.2f'))
    print()
    print('Contrasta estos totales contra el Excel antes de hacer commit.')


if __name__ == '__main__':
    main()
